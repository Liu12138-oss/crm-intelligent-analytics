from __future__ import annotations

import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_FILE = ROOT / "数据库.md"
OUTPUT_DIR = ROOT / "docs" / "db"
OVERVIEW_FILE = ROOT / "CRM数据库结构分析.md"
MODULE_FILE = ROOT / "CRM数据库业务模块梳理.md"
RELATION_FILE = ROOT / "CRM核心业务关系图.md"
XLSX_FILE = ROOT / "CRM数据库字段字典.xlsx"
MYSQL_EXE = Path(r"D:\install\mysql\bin\mysql.exe")
SYSTEM_SCHEMAS = {"information_schema", "mysql", "performance_schema", "sys"}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)
TOP_ALIGN = Alignment(vertical="top", wrap_text=True)
DEFAULT_KEY_FIELDS = [
    "id",
    "name",
    "title",
    "sn",
    "status",
    "category",
    "source",
    "stage",
    "company_name",
    "amount",
    "total_amount",
    "expect_amount",
    "organization_id",
    "user_id",
    "department_id",
    "customer_id",
    "contact_id",
    "opportunity_id",
    "contract_id",
    "order_id",
    "role_id",
    "station_id",
    "creator_id",
    "parent_id",
    "approve_status",
    "revisit_at",
    "created_at",
    "updated_at",
]
FIELD_MEANINGS = {
    "organization_id": "组织归属",
    "user_id": "负责人或所属用户",
    "department_id": "所属部门",
    "creator_id": "创建人",
    "before_user_id": "变更前负责人",
    "before_department_id": "变更前部门",
    "revisit_at": "计划回访时间",
    "real_revisit_at": "实际回访时间",
    "revisit_remind_at": "回访提醒时间",
    "approve_status": "审批状态",
    "submit_applying_at": "提交审批时间",
    "finish_approve_at": "审批完成时间",
    "created_at": "创建时间",
    "updated_at": "更新时间",
}
BUSINESS_MODULE_SPECS: list[dict[str, object]] = [
    {
        "name": "渠道/CMS与代理",
        "schema": "ikcrm_cms_production",
        "description": "承载 CRM 外围的代理商、客户套餐、试用申请、升级通知和电话池配置，更像官网/渠道运营侧库。",
        "include_all_schema_tables": True,
        "core_tables": ["agents", "agent_users", "clients"],
        "field_priority": ["id", "name", "agent_id", "status", "source", "phone", "email", "crm_user_id", "created_at", "updated_at"],
        "remarks": [
            "该模块基本覆盖整个 `ikcrm_cms_production` 库，边界相对独立。",
            "大量表围绕套餐、升级通知、试用申请展开，适合作为外围商业化配置中心。",
        ],
    },
    {
        "name": "线索管理",
        "schema": "vcooline_ikcrm_production",
        "description": "负责潜客录入、线索分配、跟进、转客户，以及线索扩展信息沉淀。",
        "focus_tables": ["leads"],
        "prefixes": ["lead_", "leads_"],
        "core_tables": ["leads"],
        "field_priority": [
            "id",
            "name",
            "company_name",
            "source",
            "status",
            "organization_id",
            "user_id",
            "department_id",
            "turned_customer_id",
            "turned_at",
            "revisit_at",
            "channel_code",
            "is_draft",
        ],
        "remarks": [
            "主表内同时保留 `turned_customer_id` 和 `turned_at`，说明线索转客户是核心流程节点。",
            "`lead_addresses`、`lead_assets`、`lead_extras` 表明线索支持地址、附件和扩展属性的柔性扩展。",
        ],
    },
    {
        "name": "客户与联系人",
        "schema": "vcooline_ikcrm_production",
        "description": "覆盖客户主数据、联系人、客户状态、回访和扩展属性，是 CRM 主数据中台的核心部分。",
        "focus_tables": ["customers", "customers_30", "contacts", "contacts_expenses", "contacts_revisit_logs"],
        "prefixes": ["customer_", "contact_"],
        "core_tables": ["customers", "contacts"],
        "field_priority": [
            "id",
            "name",
            "company_name",
            "category",
            "source",
            "industry",
            "status",
            "organization_id",
            "user_id",
            "department_id",
            "customer_common_setting_id",
            "approve_status",
            "revisit_at",
            "created_at",
            "updated_at",
        ],
        "remarks": [
            "存在 `_30` 表，说明客户相关对象可能有分表、归档或分区策略。",
            "客户模块下挂审批、通知、状态轨迹、通用设置等多类附属表，业务流程复杂度高。",
        ],
    },
    {
        "name": "商机与合同",
        "schema": "vcooline_ikcrm_production",
        "description": "承载从销售机会到合同签订的成交过程，包含阶段推进、审批和资产附件。",
        "focus_tables": ["opportunities", "contracts"],
        "prefixes": ["opportunity_", "contract_"],
        "core_tables": ["opportunities", "contracts"],
        "field_priority": [
            "id",
            "title",
            "status",
            "stage",
            "customer_id",
            "opportunity_id",
            "organization_id",
            "user_id",
            "department_id",
            "expect_amount",
            "total_amount",
            "approve_status",
            "sign_date",
            "created_at",
            "updated_at",
        ],
        "remarks": [
            "商机表按阶段推进，合同表按审批和收款金额跟踪，属于典型 B2B 销售漏斗设计。",
            "`contract_assets` 和 `opportunity_assets` 体量都较大，说明系统对附件/明细资料留存较重。",
        ],
    },
    {
        "name": "财务结算",
        "schema": "vcooline_ikcrm_production",
        "description": "覆盖回款、发票、费用、订单与支付，是 CRM 成交后的财务闭环。",
        "focus_tables": [
            "received_payments",
            "invoices",
            "invoiced_payments",
            "expenses",
            "expense_accounts",
            "orders",
            "payments",
            "bills",
            "agent_bills",
            "acceptances",
            "recharge_records",
        ],
        "prefixes": ["received_payment_", "invoice_", "expense_"],
        "core_tables": ["received_payments", "invoices", "expenses", "orders", "payments"],
        "field_priority": [
            "id",
            "sn",
            "status",
            "amount",
            "contract_id",
            "expense_account_id",
            "organization_id",
            "user_id",
            "order_id",
            "payment_type",
            "receive_date",
            "invoice_status",
            "approve_status",
            "created_at",
            "updated_at",
        ],
        "remarks": [
            "合同、回款、发票、费用之间通过应用层字段串联，而不是大量外键约束。",
            "`orders`/`payments` 更像平台通用交易层，`received_payments`/`invoices`/`expenses` 更偏 CRM 财务业务层。",
        ],
    },
    {
        "name": "组织与权限",
        "schema": "vcooline_ikcrm_production",
        "description": "提供组织、部门、用户、角色、权限与岗位等基础能力，贯穿所有核心业务实体。",
        "focus_tables": [
            "organizations",
            "departments",
            "users",
            "roles",
            "permissions",
            "grants",
            "stations",
            "ownerships",
            "common_entity_owners",
        ],
        "prefixes": ["users_", "roles_", "permissions_", "admins_"],
        "core_tables": ["organizations", "departments", "users", "roles", "permissions"],
        "field_priority": [
            "id",
            "name",
            "organization_id",
            "role_id",
            "station_id",
            "status",
            "usable",
            "user_type",
            "superior_id",
            "path",
            "created_at",
            "updated_at",
        ],
        "remarks": [
            "绝大多数业务主表都带 `organization_id`、`user_id`、`department_id`，说明数据权限是全局一等公民。",
            "`roles_users`、`permissions_roles`、`users_departments` 这些映射表是权限落地的关键对象。",
        ],
    },
    {
        "name": "通信与营销触达",
        "schema": "vcooline_ikcrm_production",
        "description": "负责呼叫、短信、社交裂变、短链、通知和销售动态触达。",
        "focus_tables": [
            "notifications",
            "reminders",
            "expire_reminders",
            "revisit_logs",
            "revisit_logs_30",
            "sales_activities",
            "sales_activity_comments",
            "sales_circles",
            "sales_circle_comments",
            "sales_circle_msgs",
            "callagents",
            "callcenters",
        ],
        "prefixes": ["call_", "dial_", "sms_", "social_", "short_link_", "ikcall_", "iksms_", "soukebox_"],
        "core_tables": ["notifications", "sales_activities", "call_records", "sms_records"],
        "field_priority": ["id", "status", "organization_id", "user_id", "customer_id", "contact_id", "created_at", "updated_at"],
        "remarks": [
            "该模块存在多套触达能力并存，包括呼叫中心、短信通道、社交分享和系统通知。",
            "历史兼容命名明显，如 `callagents`/`call_agents`、`callcenters`/`call_centers` 并存。",
        ],
    },
    {
        "name": "生态集成与平台支撑",
        "schema": "vcooline_ikcrm_production",
        "description": "覆盖钉钉、企业微信、King、阿里生态接入，以及平台配置、知识库、报表和运维元数据。",
        "focus_tables": [
            "apps",
            "app_versions",
            "subscribers",
            "api_keys",
            "settings",
            "operation_logs",
            "login_logs",
            "token_logs",
            "partitioning_configs",
            "schema_migrations",
            "sequence",
            "import_histories",
        ],
        "prefixes": [
            "ding_",
            "wx_",
            "king_",
            "alim_",
            "liteapp",
            "mina_",
            "oauth_",
            "knowledge_",
            "manual",
            "faq",
            "custom_",
            "field_",
            "data_report",
            "performance_",
            "report_",
            "organization_entity_",
        ],
        "core_tables": ["operation_logs", "token_logs", "custom_fields", "knowledge_articles"],
        "field_priority": ["id", "name", "organization_id", "user_id", "status", "created_at", "updated_at"],
        "remarks": [
            "日志、报表、自定义字段、知识库和第三方生态表共同构成平台化能力层。",
            "大表压力主要集中在该模块内的日志表和自定义字段相关表。",
        ],
    },
]
CORE_RELATIONS: list[dict[str, str]] = [
    {"group": "组织权限", "source": "departments", "column": "organization_id", "target": "organizations", "type": "从属", "description": "部门属于组织。"},
    {"group": "组织权限", "source": "users", "column": "organization_id", "target": "organizations", "type": "归属", "description": "用户归属组织。"},
    {"group": "组织权限", "source": "users", "column": "role_id", "target": "roles", "type": "默认角色", "description": "用户默认角色。"},
    {"group": "组织权限", "source": "users_departments", "column": "user_id", "target": "users", "type": "映射", "description": "用户和部门的多对多映射。"},
    {"group": "组织权限", "source": "users_departments", "column": "department_id", "target": "departments", "type": "映射", "description": "用户和部门的多对多映射。"},
    {"group": "组织权限", "source": "roles_users", "column": "role_id", "target": "roles", "type": "映射", "description": "角色用户映射。"},
    {"group": "组织权限", "source": "roles_users", "column": "user_id", "target": "users", "type": "映射", "description": "角色用户映射。"},
    {"group": "组织权限", "source": "permissions_roles", "column": "permission_id", "target": "permissions", "type": "映射", "description": "角色权限映射。"},
    {"group": "组织权限", "source": "permissions_roles", "column": "role_id", "target": "roles", "type": "映射", "description": "角色权限映射。"},
    {"group": "销售链路", "source": "leads", "column": "organization_id", "target": "organizations", "type": "归属", "description": "线索归属组织。"},
    {"group": "销售链路", "source": "leads", "column": "user_id", "target": "users", "type": "负责人", "description": "线索负责人。"},
    {"group": "销售链路", "source": "leads", "column": "department_id", "target": "departments", "type": "归属", "description": "线索归属部门。"},
    {"group": "销售链路", "source": "leads", "column": "turned_customer_id", "target": "customers", "type": "转化", "description": "线索转客户后的目标客户。"},
    {"group": "销售链路", "source": "contacts", "column": "customer_id", "target": "customers", "type": "从属", "description": "联系人挂靠客户。"},
    {"group": "销售链路", "source": "customers", "column": "organization_id", "target": "organizations", "type": "归属", "description": "客户归属组织。"},
    {"group": "销售链路", "source": "customers", "column": "user_id", "target": "users", "type": "负责人", "description": "客户负责人。"},
    {"group": "销售链路", "source": "customers", "column": "department_id", "target": "departments", "type": "归属", "description": "客户归属部门。"},
    {"group": "销售链路", "source": "opportunities", "column": "customer_id", "target": "customers", "type": "来源客户", "description": "商机挂靠客户。"},
    {"group": "销售链路", "source": "opportunities", "column": "organization_id", "target": "organizations", "type": "归属", "description": "商机归属组织。"},
    {"group": "销售链路", "source": "opportunities", "column": "user_id", "target": "users", "type": "负责人", "description": "商机负责人。"},
    {"group": "销售链路", "source": "opportunities", "column": "department_id", "target": "departments", "type": "归属", "description": "商机归属部门。"},
    {"group": "销售链路", "source": "contracts", "column": "customer_id", "target": "customers", "type": "签约客户", "description": "合同对应客户。"},
    {"group": "销售链路", "source": "contracts", "column": "opportunity_id", "target": "opportunities", "type": "来源商机", "description": "合同通常由商机推进而来。"},
    {"group": "销售链路", "source": "contracts", "column": "organization_id", "target": "organizations", "type": "归属", "description": "合同归属组织。"},
    {"group": "销售链路", "source": "contracts", "column": "user_id", "target": "users", "type": "负责人", "description": "合同负责人。"},
    {"group": "销售链路", "source": "contracts", "column": "department_id", "target": "departments", "type": "归属", "description": "合同归属部门。"},
    {"group": "财务闭环", "source": "received_payment_plans", "column": "contract_id", "target": "contracts", "type": "计划归属", "description": "回款计划属于合同。"},
    {"group": "财务闭环", "source": "received_payments", "column": "contract_id", "target": "contracts", "type": "实际回款", "description": "实际回款归属于合同。"},
    {"group": "财务闭环", "source": "received_payments", "column": "received_payment_plan_id", "target": "received_payment_plans", "type": "落计划", "description": "实际回款可对应一条回款计划。"},
    {"group": "财务闭环", "source": "received_payments", "column": "organization_id", "target": "organizations", "type": "归属", "description": "回款记录归属组织。"},
    {"group": "财务闭环", "source": "received_payments", "column": "user_id", "target": "users", "type": "负责人", "description": "回款记录负责人。"},
    {"group": "财务闭环", "source": "invoiced_payments", "column": "contract_id", "target": "contracts", "type": "开票记录", "description": "合同维度的已开票记录。"},
    {"group": "财务闭环", "source": "invoice_items", "column": "invoice_id", "target": "invoices", "type": "明细归属", "description": "发票项目明细归属于发票主表。"},
    {"group": "财务闭环", "source": "invoices", "column": "organization_id", "target": "organizations", "type": "归属", "description": "发票申请归属组织。"},
    {"group": "财务闭环", "source": "invoices", "column": "user_id", "target": "users", "type": "申请人/负责人", "description": "发票申请用户。"},
    {"group": "财务闭环", "source": "expense_accounts", "column": "organization_id", "target": "organizations", "type": "归属", "description": "费用台账归属组织。"},
    {"group": "财务闭环", "source": "expense_accounts", "column": "department_id", "target": "departments", "type": "归属", "description": "费用台账归属部门。"},
    {"group": "财务闭环", "source": "expense_accounts", "column": "user_id", "target": "users", "type": "负责人", "description": "费用台账负责人。"},
    {"group": "财务闭环", "source": "expenses", "column": "expense_account_id", "target": "expense_accounts", "type": "费用归集", "description": "费用记录归入费用台账。"},
    {"group": "财务闭环", "source": "expenses", "column": "customer_id", "target": "customers", "type": "关联客户", "description": "费用可归属到客户。"},
    {"group": "财务闭环", "source": "expenses", "column": "organization_id", "target": "organizations", "type": "归属", "description": "费用记录归属组织。"},
    {"group": "财务闭环", "source": "expenses", "column": "user_id", "target": "users", "type": "负责人", "description": "费用记录负责人。"},
    {"group": "财务闭环", "source": "payments", "column": "order_id", "target": "orders", "type": "支付对应订单", "description": "支付记录关联平台订单。"},
    {"group": "财务闭环", "source": "payments", "column": "organization_id", "target": "organizations", "type": "归属", "description": "支付记录归属组织。"},
    {"group": "财务闭环", "source": "orders", "column": "organization_id", "target": "organizations", "type": "归属", "description": "订单归属组织。"},
]


@dataclass
class DbConfig:
    host: str
    port: str
    user: str
    password: str


def parse_db_config(path: Path) -> DbConfig:
    text = path.read_text(encoding="utf-8")
    items: dict[str, str] = {}
    for line in text.splitlines():
        if "：" not in line:
            continue
        key, value = [part.strip() for part in line.split("：", 1)]
        items[key] = value
    return DbConfig(
        host=items["主机地址"],
        port=items["端口"],
        user=items["用户名"],
        password=items["密码"],
    )


def mysql_query(config: DbConfig, sql: str) -> list[list[str]]:
    cmd = [
        str(MYSQL_EXE),
        f"--host={config.host}",
        f"--port={config.port}",
        f"--user={config.user}",
        f"--password={config.password}",
        "--default-character-set=utf8mb4",
        "-N",
        "-B",
        "-e",
        sql,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace", check=True)
    rows: list[list[str]] = []
    for line in result.stdout.splitlines():
        if not line.strip():
            continue
        rows.append(line.split("\t"))
    return rows


def sql_quote(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "\\'") + "'"


def load_metadata(config: DbConfig) -> dict[str, object]:
    schema_rows = mysql_query(config, "SHOW DATABASES;")
    schemas = [row[0] for row in schema_rows if row[0] not in SYSTEM_SCHEMAS]
    schema_list = ", ".join(sql_quote(name) for name in schemas)

    table_rows = mysql_query(
        config,
        f"""
        SELECT
            table_schema,
            table_name,
            engine,
            IFNULL(table_rows, 0),
            IFNULL(data_length, 0),
            IFNULL(index_length, 0),
            IFNULL(create_time, ''),
            IFNULL(update_time, ''),
            IFNULL(table_comment, '')
        FROM information_schema.tables
        WHERE table_schema IN ({schema_list})
          AND table_type = 'BASE TABLE'
        ORDER BY table_schema, table_name;
        """,
    )
    column_rows = mysql_query(
        config,
        f"""
        SELECT
            table_schema,
            table_name,
            ordinal_position,
            column_name,
            column_type,
            is_nullable,
            IFNULL(column_default, 'NULL'),
            column_key,
            extra,
            IFNULL(column_comment, '')
        FROM information_schema.columns
        WHERE table_schema IN ({schema_list})
        ORDER BY table_schema, table_name, ordinal_position;
        """,
    )
    index_rows = mysql_query(
        config,
        f"""
        SELECT
            table_schema,
            table_name,
            index_name,
            non_unique,
            seq_in_index,
            column_name,
            IFNULL(collation, ''),
            IFNULL(sub_part, ''),
            index_type
        FROM information_schema.statistics
        WHERE table_schema IN ({schema_list})
        ORDER BY table_schema, table_name, index_name, seq_in_index;
        """,
    )
    foreign_key_rows = mysql_query(
        config,
        f"""
        SELECT
            table_schema,
            table_name,
            column_name,
            referenced_table_name,
            referenced_column_name,
            constraint_name
        FROM information_schema.key_column_usage
        WHERE table_schema IN ({schema_list})
          AND referenced_table_name IS NOT NULL
        ORDER BY table_schema, table_name, constraint_name, ordinal_position;
        """,
    )

    tables: dict[str, dict[str, dict[str, object]]] = defaultdict(dict)
    for row in table_rows:
        schema, table_name, engine, table_rows_value, data_length, index_length, create_time, update_time, comment = row
        tables[schema][table_name] = {
            "schema": schema,
            "table_name": table_name,
            "engine": engine,
            "table_rows": int(table_rows_value),
            "data_length": int(data_length),
            "index_length": int(index_length),
            "create_time": create_time,
            "update_time": update_time,
            "comment": comment,
            "columns": [],
            "indexes": defaultdict(list),
            "foreign_keys": [],
        }

    for row in column_rows:
        schema, table_name, position, column_name, column_type, nullable, default, column_key, extra, comment = row
        tables[schema][table_name]["columns"].append(
            {
                "position": int(position),
                "name": column_name,
                "type": column_type,
                "nullable": nullable,
                "default": default,
                "key": column_key,
                "extra": extra,
                "comment": comment,
            }
        )

    for row in index_rows:
        schema, table_name, index_name, non_unique, seq_in_index, column_name, collation, sub_part, index_type = row
        tables[schema][table_name]["indexes"][index_name].append(
            {
                "non_unique": int(non_unique),
                "seq_in_index": int(seq_in_index),
                "column_name": column_name,
                "collation": collation,
                "sub_part": sub_part,
                "index_type": index_type,
            }
        )

    for row in foreign_key_rows:
        schema, table_name, column_name, ref_table, ref_column, constraint_name = row
        tables[schema][table_name]["foreign_keys"].append(
            {
                "constraint_name": constraint_name,
                "column_name": column_name,
                "referenced_table_name": ref_table,
                "referenced_column_name": ref_column,
            }
        )

    return {
        "schemas": schemas,
        "tables": tables,
    }


def format_size(num_bytes: int) -> str:
    if num_bytes <= 0:
        return "0 MB"
    mb = num_bytes / 1024 / 1024
    if mb >= 1024:
        return f"{mb / 1024:.2f} GB"
    return f"{mb:.2f} MB"


def infer_module_name(table_name: str) -> str:
    if table_name.startswith(("ding_", "wx_", "king_", "alim_", "liteapp_", "mina_", "wechat_")):
        return "生态集成"
    if table_name.startswith(("call_", "sms_", "social_", "short_link_", "dial_", "ikcall_", "soukebox_")):
        return "通信与触达"
    if table_name.startswith(("knowledge_", "manual", "faq", "cms_", "announcement", "feedback")):
        return "内容与知识"
    if table_name.startswith(("expense", "invoice", "payment", "received_", "commission", "payslip", "bill", "recharge_")):
        return "财务结算"
    if table_name.startswith(("report", "data_report", "schedule_report", "performance_", "organization_entity_")):
        return "报表分析"
    if table_name.startswith(("custom_", "field_", "tag", "ik_tag", "marking", "print_template", "settings", "partitioning_")):
        return "平台配置"
    if table_name.startswith(("approval", "approve", "multistep_", "notify_", "reminder", "notification", "event", "checkin")):
        return "流程协同"
    if table_name.startswith(("operation_log", "login_log", "token_log", "import_histories", "schema_migrations", "sequence")):
        return "日志与运维"
    if table_name.startswith(("agent_", "client", "supplier_", "package", "notice", "phone_pool", "kebao_", "purchase_")):
        return "渠道与代理"
    crm_core_prefixes = (
        "customer",
        "contact",
        "lead",
        "opportunity",
        "contract",
        "order",
        "product",
        "account",
        "address",
        "asset",
        "ownership",
        "station",
        "organization",
        "department",
        "role",
        "permission",
        "user",
        "admin",
        "customer_",
        "contact_",
        "lead_",
        "opportunity_",
        "contract_",
        "product_",
    )
    if table_name.startswith(crm_core_prefixes) or table_name in {
        "customers",
        "contacts",
        "leads",
        "opportunities",
        "contracts",
        "orders",
        "products",
        "users",
        "roles",
        "permissions",
        "departments",
        "organizations",
        "accounts",
        "addresses",
        "expenses",
    }:
        return "CRM核心业务"
    return "其他支撑"


def collect_summary(metadata: dict[str, object]) -> dict[str, dict[str, object]]:
    tables = metadata["tables"]  # type: ignore[assignment]
    summaries: dict[str, dict[str, object]] = {}
    for schema, schema_tables in tables.items():  # type: ignore[union-attr]
        column_count = 0
        indexes_count = 0
        foreign_key_count = 0
        total_size = 0
        modules = Counter()
        largest_tables: list[dict[str, object]] = []
        no_comment_tables = 0

        for table in schema_tables.values():
            column_count += len(table["columns"])
            indexes_count += len(table["indexes"])
            foreign_key_count += len(table["foreign_keys"])
            table_size = table["data_length"] + table["index_length"]
            total_size += table_size
            modules[infer_module_name(table["table_name"])] += 1
            if not table["comment"]:
                no_comment_tables += 1
            largest_tables.append(
                {
                    "table_name": table["table_name"],
                    "rows": table["table_rows"],
                    "size_bytes": table_size,
                }
            )

        largest_tables.sort(key=lambda item: (item["size_bytes"], item["rows"]), reverse=True)
        summaries[schema] = {
            "table_count": len(schema_tables),
            "column_count": column_count,
            "indexes_count": indexes_count,
            "foreign_key_count": foreign_key_count,
            "total_size": total_size,
            "modules": modules,
            "largest_tables": largest_tables[:10],
            "no_comment_tables": no_comment_tables,
        }
    return summaries


def pluralize(noun: str) -> str:
    if noun.endswith("y") and len(noun) > 1 and noun[-2] not in "aeiou":
        return noun[:-1] + "ies"
    if noun.endswith(("s", "x", "z", "ch", "sh")):
        return noun + "es"
    return noun + "s"


def infer_target_candidates(column_name: str) -> list[str]:
    if not column_name.endswith("_id"):
        return []
    base = column_name[:-3]
    candidates = [base, pluralize(base)]
    unique_candidates: list[str] = []
    for candidate in candidates:
        if candidate not in unique_candidates:
            unique_candidates.append(candidate)
    return unique_candidates


def build_reference_indexes(schema_tables: dict[str, dict[str, object]]) -> tuple[dict[str, list[dict[str, str]]], dict[str, list[dict[str, str]]]]:
    table_names = set(schema_tables.keys())
    forward_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    reverse_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    seen_pairs: set[tuple[str, str, str]] = set()

    for table in schema_tables.values():
        for column in table["columns"]:
            for candidate in infer_target_candidates(column["name"]):
                if candidate not in table_names or candidate == table["table_name"]:
                    continue
                pair = (table["table_name"], column["name"], candidate)
                if pair in seen_pairs:
                    continue
                seen_pairs.add(pair)
                relation = {
                    "source_table": table["table_name"],
                    "column_name": column["name"],
                    "target_table": candidate,
                }
                forward_index[table["table_name"]].append(relation)
                reverse_index[candidate].append(relation)
                break

    return forward_index, reverse_index


def resolve_module_tables(spec: dict[str, object], schema_tables: dict[str, dict[str, object]]) -> list[dict[str, object]]:
    if spec.get("include_all_schema_tables"):
        return sorted(schema_tables.values(), key=lambda item: item["table_name"])

    selected: dict[str, dict[str, object]] = {}
    for table_name in spec.get("focus_tables", []):
        if table_name in schema_tables:
            selected[table_name] = schema_tables[table_name]
    for prefix in spec.get("prefixes", []):
        for table_name, table in schema_tables.items():
            if table_name.startswith(prefix):
                selected[table_name] = table
    return sorted(selected.values(), key=lambda item: item["table_name"])


def pick_key_fields(table: dict[str, object], preferred_fields: list[str], limit: int = 12) -> list[str]:
    column_map = {column["name"]: column["type"] for column in table["columns"]}
    ordered_fields: list[str] = []
    for field_name in [*preferred_fields, *DEFAULT_KEY_FIELDS, *column_map.keys()]:
        if field_name in column_map and field_name not in ordered_fields:
            ordered_fields.append(field_name)
        if len(ordered_fields) >= limit:
            break
    return [f"`{field}` (`{column_map[field]}`)" for field in ordered_fields]


def summarize_table_list(table_names: list[str], limit: int = 12) -> str:
    if not table_names:
        return "-"
    preview = ", ".join(f"`{name}`" for name in table_names[:limit])
    if len(table_names) > limit:
        preview += f" 等 {len(table_names)} 张表"
    return preview


def format_relations(relations: list[dict[str, str]], mode: str, limit: int = 6) -> str:
    if not relations:
        return "未从字段命名中推断到稳定关系。"

    relation_texts: list[str] = []
    for relation in relations[:limit]:
        if mode == "forward":
            relation_texts.append(f"`{relation['column_name']}` -> `{relation['target_table']}`")
        else:
            relation_texts.append(f"`{relation['source_table']}.{relation['column_name']}`")
    suffix = ""
    if len(relations) > limit:
        suffix = f" 等 {len(relations)} 条"
    return "、".join(relation_texts) + suffix


def collect_common_field_rows(metadata: dict[str, object]) -> list[tuple[str, int, str]]:
    tables: dict[str, dict[str, dict[str, object]]] = metadata["tables"]  # type: ignore[assignment]
    core_table_pairs: list[tuple[str, str]] = []
    for spec in BUSINESS_MODULE_SPECS:
        schema = str(spec["schema"])
        for table_name in spec.get("core_tables", []):
            if table_name in tables.get(schema, {}):
                core_table_pairs.append((schema, table_name))

    field_counter: Counter[str] = Counter()
    for schema, table_name in core_table_pairs:
        column_names = {column["name"] for column in tables[schema][table_name]["columns"]}
        for column_name in column_names:
            field_counter[column_name] += 1

    common_rows: list[tuple[str, int, str]] = []
    for field_name, count in field_counter.most_common():
        if count < 3:
            continue
        meaning = FIELD_MEANINGS.get(field_name, "-")
        common_rows.append((field_name, count, meaning))
        if len(common_rows) >= 12:
            break
    return common_rows


def collect_module_rows(metadata: dict[str, object]) -> list[dict[str, object]]:
    tables: dict[str, dict[str, dict[str, object]]] = metadata["tables"]  # type: ignore[assignment]
    module_rows: list[dict[str, object]] = []
    for spec in BUSINESS_MODULE_SPECS:
        schema = str(spec["schema"])
        schema_tables = tables[schema]
        selected_tables = resolve_module_tables(spec, schema_tables)
        table_names = [table["table_name"] for table in selected_tables]
        total_rows = sum(int(table["table_rows"]) for table in selected_tables)
        total_size = sum(int(table["data_length"]) + int(table["index_length"]) for table in selected_tables)
        core_tables = [schema_tables[table_name] for table_name in spec.get("core_tables", []) if table_name in schema_tables]
        module_rows.append(
            {
                "spec": spec,
                "tables": selected_tables,
                "table_names": table_names,
                "total_rows": total_rows,
                "total_size": total_size,
                "core_tables": core_tables,
            }
        )
    return module_rows


def build_module_lookup(module_rows: list[dict[str, object]]) -> dict[tuple[str, str], str]:
    lookup: dict[tuple[str, str], str] = {}
    for module in module_rows:
        spec = module["spec"]
        schema = str(spec["schema"])
        module_name = str(spec["name"])
        for table_name in module["table_names"]:
            lookup[(schema, table_name)] = module_name
    return lookup


def build_module_markdown(metadata: dict[str, object], summaries: dict[str, dict[str, object]]) -> str:
    tables: dict[str, dict[str, dict[str, object]]] = metadata["tables"]  # type: ignore[assignment]
    reference_indexes = {
        schema: build_reference_indexes(schema_tables)
        for schema, schema_tables in tables.items()
    }
    module_rows = collect_module_rows(metadata)

    lines: list[str] = []
    lines.append("# CRM数据库业务模块梳理")
    lines.append("")
    lines.append("## 1. 文档目标")
    lines.append("")
    lines.append("- 这份文档不是逐表字典，而是把数据库按业务域重新组织，方便产品、开发、实施和数据同学快速理解系统边界。")
    lines.append("- 逐表字段、索引、外键明细仍以 `CRM数据库结构分析.md` 和 `docs/db/*.md` 为准。")
    lines.append("- 模块归类基于表名前缀、字段模式和核心实体关系推断，适合作为数据库阅读地图。")
    lines.append("")
    lines.append("## 2. 核心业务链路")
    lines.append("")
    lines.append("1. `ikcrm_cms_production` 负责渠道代理、套餐和试用入口，为核心 CRM 输送外围客户与配置。")
    lines.append("2. `leads` 承接线索录入和分配，通过 `turned_customer_id` 将有效线索转成客户。")
    lines.append("3. `customers`/`contacts` 形成客户主数据层，承接回访、状态轨迹、扩展属性和联系人关系。")
    lines.append("4. `opportunities` 跟踪成交机会，`contracts` 负责正式签约。")
    lines.append("5. `received_payments`、`invoices`、`expenses` 形成签约后的财务闭环；`orders`、`payments` 则承担平台通用交易能力。")
    lines.append("6. `organizations`、`departments`、`users`、`roles`、`permissions` 贯穿所有业务域，构成组织和权限底座。")
    lines.append("")
    lines.append("## 3. 模块总览")
    lines.append("")
    lines.append("| 模块 | 所属库 | 表数量 | 预估行数 | 预估容量 | 核心表 |")
    lines.append("| --- | --- | ---: | ---: | ---: | --- |")
    for module in module_rows:
        spec = module["spec"]
        core_table_names = [table["table_name"] for table in module["core_tables"]]
        lines.append(
            f"| {spec['name']} | `{spec['schema']}` | {len(module['tables'])} | {module['total_rows']} | "
            f"{format_size(module['total_size'])} | {summarize_table_list(core_table_names, limit=6)} |"
        )
    lines.append("")
    lines.append("## 4. 核心实体通用字段模式")
    lines.append("")
    lines.append("| 字段 | 出现于核心表数量 | 典型含义 |")
    lines.append("| --- | ---: | --- |")
    for field_name, count, meaning in collect_common_field_rows(metadata):
        lines.append(f"| `{field_name}` | {count} | {meaning} |")
    lines.append("")
    lines.append("## 5. 模块详解")
    lines.append("")

    for index, module in enumerate(module_rows, start=1):
        spec = module["spec"]
        schema = str(spec["schema"])
        forward_index, reverse_index = reference_indexes[schema]
        companion_table_names = [name for name in module["table_names"] if name not in [table["table_name"] for table in module["core_tables"]]]
        largest_tables = sorted(
            module["tables"],
            key=lambda item: (item["data_length"] + item["index_length"], item["table_rows"]),
            reverse=True,
        )[:5]

        lines.append(f"### 5.{index} {spec['name']}")
        lines.append("")
        lines.append(f"- 模块定位：{spec['description']}")
        lines.append(
            f"- 范围统计：覆盖 `{schema}` 中 {len(module['tables'])} 张表，预估总行数 {module['total_rows']}，预估总容量 {format_size(module['total_size'])}。"
        )
        lines.append(f"- 核心表：{summarize_table_list([table['table_name'] for table in module['core_tables']], limit=8)}")
        lines.append(f"- 附属/映射表：{summarize_table_list(companion_table_names, limit=10)}")
        if largest_tables:
            lines.append(
                "- 模块内较大的表："
                + "、".join(
                    f"`{table['table_name']}` ({table['table_rows']} 行, {format_size(table['data_length'] + table['index_length'])})"
                    for table in largest_tables
                )
            )
        lines.append("")
        lines.append("| 主表 | 关键字段 | 该表主动关联对象 | 其他表对它的依赖 |")
        lines.append("| --- | --- | --- | --- |")
        for table in module["core_tables"]:
            lines.append(
                f"| `{table['table_name']}` | {', '.join(pick_key_fields(table, spec.get('field_priority', []), limit=12))} | "
                f"{format_relations(forward_index.get(table['table_name'], []), mode='forward')} | "
                f"{format_relations(reverse_index.get(table['table_name'], []), mode='reverse')} |"
            )
        lines.append("")
        for remark in spec.get("remarks", []):
            lines.append(f"- 观察：{remark}")
        lines.append("")

    lines.append("## 6. 阅读建议")
    lines.append("")
    lines.append("1. 先看本文件的业务链路和模块详解，再去 `docs/db/*.md` 查单表字段。")
    lines.append("2. 对没有显式外键的关系，优先关注 `_id` 字段、映射表和审批/通知附表。")
    lines.append("3. 如果后续要做数据治理，建议先从 `组织与权限`、`客户与联系人`、`商机与合同` 三个模块补注释和口径。")
    lines.append("")
    return "\n".join(lines) + "\n"


def get_table(metadata: dict[str, object], schema: str, table_name: str) -> dict[str, object] | None:
    tables: dict[str, dict[str, dict[str, object]]] = metadata["tables"]  # type: ignore[assignment]
    return tables.get(schema, {}).get(table_name)


def build_sales_mermaid() -> str:
    lines = [
        "```mermaid",
        "flowchart LR",
        "    leads[线索 leads]",
        "    customers[客户 customers]",
        "    contacts[联系人 contacts]",
        "    opportunities[商机 opportunities]",
        "    contracts[合同 contracts]",
        "    plans[回款计划 received_payment_plans]",
        "    receipts[回款 received_payments]",
        "    invoiced[已开票记录 invoiced_payments]",
        "    invoices[发票 invoices]",
        "    expenses[费用 expenses]",
        "    expense_accounts[费用台账 expense_accounts]",
        "    leads -->|turned_customer_id| customers",
        "    customers -->|customer_id| contacts",
        "    customers -->|customer_id| opportunities",
        "    opportunities -->|opportunity_id| contracts",
        "    customers -->|customer_id| contracts",
        "    contracts -->|contract_id| plans",
        "    plans -->|received_payment_plan_id| receipts",
        "    contracts -->|contract_id| receipts",
        "    contracts -->|contract_id| invoiced",
        "    invoices -->|invoice_id| invoice_items[发票项 invoice_items]",
        "    expense_accounts -->|expense_account_id| expenses",
        "    customers -->|customer_id| expenses",
        "```",
    ]
    return "\n".join(lines)


def build_org_mermaid() -> str:
    lines = [
        "```mermaid",
        "flowchart LR",
        "    organizations[组织 organizations]",
        "    departments[部门 departments]",
        "    users[用户 users]",
        "    roles[角色 roles]",
        "    permissions[权限 permissions]",
        "    users_departments[users_departments]",
        "    roles_users[roles_users]",
        "    permissions_roles[permissions_roles]",
        "    organizations -->|organization_id| departments",
        "    organizations -->|organization_id| users",
        "    roles -->|role_id| users",
        "    users -->|user_id| users_departments",
        "    departments -->|department_id| users_departments",
        "    roles -->|role_id| roles_users",
        "    users -->|user_id| roles_users",
        "    roles -->|role_id| permissions_roles",
        "    permissions -->|permission_id| permissions_roles",
        "```",
    ]
    return "\n".join(lines)


def build_table_stat_row(table: dict[str, object], key_fields: list[str]) -> str:
    fields = ", ".join(pick_key_fields(table, key_fields, limit=10))
    size = format_size(table["data_length"] + table["index_length"])
    return f"| `{table['table_name']}` | {table['table_rows']} | {size} | {fields} |"


def build_relation_markdown(metadata: dict[str, object]) -> str:
    schema = "vcooline_ikcrm_production"
    key_field_map = {
        "organizations": ["id", "name", "user_id", "contacts_count", "customers_count", "leads_count", "users_count", "activity_at"],
        "departments": ["id", "name", "organization_id", "parent_id", "path", "status"],
        "users": ["id", "name", "organization_id", "role_id", "station_id", "superior_id", "user_type", "status", "usable"],
        "roles": ["id", "name", "organization_id", "entity_grant_scope", "field_permission_grant_scope"],
        "permissions": ["id", "name", "subject", "action"],
        "leads": ["id", "name", "company_name", "source", "status", "organization_id", "user_id", "department_id", "turned_customer_id", "turned_at"],
        "customers": ["id", "name", "company_name", "category", "source", "status", "organization_id", "user_id", "department_id", "customer_common_setting_id", "approve_status"],
        "contacts": ["id", "name", "customer_id", "organization_id", "user_id", "job", "category"],
        "opportunities": ["id", "title", "customer_id", "stage", "expect_amount", "organization_id", "user_id", "department_id", "approve_status"],
        "contracts": ["id", "title", "customer_id", "opportunity_id", "status", "total_amount", "organization_id", "user_id", "department_id", "approve_status"],
        "received_payment_plans": ["id", "contract_id", "receive_stage", "receive_date", "amount", "received_amount", "status"],
        "received_payments": ["id", "contract_id", "received_payment_plan_id", "amount", "payment_type", "receive_date", "invoice_status", "approve_status"],
        "invoices": ["id", "organization_id", "user_id", "invoice_type", "amount", "company_name", "check_status"],
        "invoiced_payments": ["id", "contract_id", "user_id", "invoice_types", "amount", "invoiced_date", "invoice_no"],
        "invoice_items": ["id", "invoice_id", "itemable_id", "itemable_type"],
        "expense_accounts": ["id", "organization_id", "department_id", "user_id", "sn", "amount", "approve_status"],
        "expenses": ["id", "expense_account_id", "customer_id", "organization_id", "user_id", "sn", "category", "amount", "expense_status"],
        "orders": ["id", "organization_id", "amount", "currency", "order_number", "status", "orderable_id", "orderable_type"],
        "payments": ["id", "organization_id", "order_id", "amount", "status", "channel", "transaction_number"],
    }
    focus_tables = [
        "organizations",
        "departments",
        "users",
        "roles",
        "permissions",
        "leads",
        "customers",
        "contacts",
        "opportunities",
        "contracts",
        "received_payment_plans",
        "received_payments",
        "invoices",
        "invoiced_payments",
        "invoice_items",
        "expense_accounts",
        "expenses",
        "orders",
        "payments",
    ]

    lines: list[str] = []
    lines.append("# CRM核心业务关系图")
    lines.append("")
    lines.append("## 1. 说明")
    lines.append("")
    lines.append("- 这份文档聚焦 `vcooline_ikcrm_production` 的核心业务对象关系，方便理解 CRM 主流程。")
    lines.append("- 图中关系主要根据字段命名和业务语义推断，并不等同于数据库中真实存在的外键约束。")
    lines.append("- 目前主业务库显式外键只有 3 条，因此阅读时应把它视为“业务关系图”，而不是严格的物理外键图。")
    lines.append("")
    lines.append("## 2. 销售到回款主链路")
    lines.append("")
    lines.append(build_sales_mermaid())
    lines.append("")
    lines.append("### 2.1 主链路解读")
    lines.append("")
    lines.append("1. `leads` 保存潜在线索，通过 `turned_customer_id` 记录转客户结果。")
    lines.append("2. `customers` 是核心客户主数据，`contacts` 作为客户联系人从属于客户。")
    lines.append("3. `opportunities` 由客户沉淀出的销售机会，`contracts` 则是成交后正式签约对象。")
    lines.append("4. `received_payment_plans` 和 `received_payments` 分别对应计划回款与实际回款。")
    lines.append("5. `invoices`/`invoice_items` 管理发票申请及条目，`invoiced_payments` 则反映合同维度的已开票记录。")
    lines.append("6. `expense_accounts`/`expenses` 构成费用核算链路，与客户、部门、用户有交叉关联。")
    lines.append("")
    lines.append("## 3. 组织与权限底座")
    lines.append("")
    lines.append(build_org_mermaid())
    lines.append("")
    lines.append("### 3.1 底座解读")
    lines.append("")
    lines.append("1. `organizations`、`departments`、`users` 是所有业务对象的统一归属维度。")
    lines.append("2. `users.role_id` 提供用户默认角色，同时 `roles_users` 支持多角色映射。")
    lines.append("3. `permissions_roles` 维护角色和权限的分配，是权限生效的关键枢纽。")
    lines.append("4. 多数业务主表都带 `organization_id`、`user_id`、`department_id`，说明组织和数据权限设计是全局性的。")
    lines.append("")
    lines.append("## 4. 核心表速查")
    lines.append("")
    lines.append("| 表名 | 预估行数 | 预估容量 | 关键字段 |")
    lines.append("| --- | ---: | ---: | --- |")
    for table_name in focus_tables:
        table = get_table(metadata, schema, table_name)
        if not table:
            continue
        lines.append(build_table_stat_row(table, key_field_map.get(table_name, [])))
    lines.append("")
    lines.append("## 5. 关键关系矩阵")
    lines.append("")
    lines.append("| 关系组 | 来源表 | 关系字段 | 目标表 | 关系类型 | 说明 |")
    lines.append("| --- | --- | --- | --- | --- | --- |")
    for relation in CORE_RELATIONS:
        lines.append(
            f"| {relation['group']} | `{relation['source']}` | `{relation['column']}` | "
            f"`{relation['target']}` | {relation['type']} | {relation['description']} |"
        )
    lines.append("")
    lines.append("## 6. 常见结构模式")
    lines.append("")
    lines.append("### 6.1 主表 + 资产扩展表")
    lines.append("")
    lines.append("- `lead_assets`、`customer_assets`、`opportunity_assets`、`contract_assets`、`received_payment_assets`、`expense_assets` 的结构非常相似。")
    lines.append("- 这些表通常通过 `entity_id` + `custom_field_id` 关联业务对象和自定义字段，说明系统采用了统一的动态字段扩展模型。")
    lines.append("")
    lines.append("### 6.2 主表 + 审批表")
    lines.append("")
    lines.append("- `customer_multistep_approves`、`opportunity_multistep_approves`、`contract_multistep_approves`、`expense_account_multistep_approves` 都体现了分步审批机制。")
    lines.append("- 对应主表里普遍存在 `approve_status`、`step`、`pending_step`、`submit_applying_at`、`finish_approve_at` 等字段。")
    lines.append("")
    lines.append("### 6.3 主表 + 通知映射表")
    lines.append("")
    lines.append("- `customer_notify_user_maps`、`opportunity_notify_user_maps`、`contract_notify_user_maps`、`received_payment_notify_user_maps` 说明提醒/抄送是独立建模的。")
    lines.append("- 这种设计便于一条业务记录挂多个通知用户，也减少了主表字段膨胀。")
    lines.append("")
    lines.append("### 6.4 地址与多态结构")
    lines.append("")
    lines.append("- `lead_addresses`、`customer_addresses` 等地址表使用 `addressable_id` + `addressable_type`，属于多态关联设计。")
    lines.append("- `orders.orderable_id/orderable_type`、`invoice_items.itemable_id/itemable_type`、`payments.object_id` 也体现出平台层对多业务对象复用一套交易模型。")
    lines.append("")
    lines.append("## 7. 阅读这套库的建议顺序")
    lines.append("")
    lines.append("1. 先读 `organizations`、`departments`、`users`、`roles`、`permissions`，理解数据归属和权限边界。")
    lines.append("2. 再读 `leads`、`customers`、`contacts`、`opportunities`、`contracts`，掌握销售转化主链路。")
    lines.append("3. 最后读 `received_payments`、`invoices`、`expenses`、`orders`、`payments`，理解财务与平台交易层的差异。")
    lines.append("")
    return "\n".join(lines) + "\n"


def style_sheet(ws, freeze_cell: str, autofilter: bool = True) -> None:
    ws.freeze_panes = freeze_cell
    if autofilter and ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = TOP_ALIGN
            cell.border = THIN_BORDER


def apply_header_style(ws, row_idx: int, fill: PatternFill, font: Font) -> None:
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = TOP_ALIGN
        cell.border = THIN_BORDER


def autosize_columns(ws, min_width: int = 10, max_width: int = 60) -> None:
    widths: dict[int, int] = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell_len = len(str(cell.value))
            widths[cell.column] = min(max(widths[cell.column], cell_len + 2), max_width)
    for column_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(column_idx)].width = max(width, min_width)


def append_sheet_rows(ws, headers: list[str], rows: list[list[object]], title: str | None = None) -> None:
    current_row = 1
    if title:
        ws.cell(row=current_row, column=1, value=title)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max(len(headers), 1))
        apply_header_style(ws, current_row, SUBHEADER_FILL, SUBHEADER_FONT)
        current_row += 1
    ws.append(headers)
    apply_header_style(ws, current_row, HEADER_FILL, HEADER_FONT)
    for row in rows:
        ws.append(row)
    style_sheet(ws, freeze_cell=f"A{current_row + 1}")
    autosize_columns(ws)


def build_overview_rows(metadata: dict[str, object], summaries: dict[str, dict[str, object]]) -> list[list[object]]:
    rows: list[list[object]] = []
    for schema in metadata["schemas"]:
        summary = summaries[schema]
        no_comment_ratio = 0.0
        if summary["table_count"]:
            no_comment_ratio = summary["no_comment_tables"] / summary["table_count"] * 100
        rows.append(
            [
                schema,
                summary["table_count"],
                summary["column_count"],
                summary["indexes_count"],
                summary["foreign_key_count"],
                format_size(summary["total_size"]),
                f"{no_comment_ratio:.1f}%",
            ]
        )
    return rows


def build_table_rows(
    metadata: dict[str, object],
    module_lookup: dict[tuple[str, str], str],
) -> list[list[object]]:
    tables: dict[str, dict[str, dict[str, object]]] = metadata["tables"]  # type: ignore[assignment]
    rows: list[list[object]] = []
    for schema in metadata["schemas"]:
        for table in sorted(tables[schema].values(), key=lambda item: item["table_name"]):
            rows.append(
                [
                    schema,
                    table["table_name"],
                    module_lookup.get((schema, table["table_name"]), infer_module_name(table["table_name"])),
                    table["engine"],
                    len(table["columns"]),
                    len(table["indexes"]),
                    len(table["foreign_keys"]),
                    table["table_rows"],
                    format_size(table["data_length"]),
                    format_size(table["index_length"]),
                    format_size(table["data_length"] + table["index_length"]),
                    table["create_time"] or "",
                    table["update_time"] or "",
                    table["comment"] or "",
                ]
            )
    return rows


def build_column_rows(
    metadata: dict[str, object],
    module_lookup: dict[tuple[str, str], str],
) -> list[list[object]]:
    tables: dict[str, dict[str, dict[str, object]]] = metadata["tables"]  # type: ignore[assignment]
    rows: list[list[object]] = []
    for schema in metadata["schemas"]:
        for table in sorted(tables[schema].values(), key=lambda item: item["table_name"]):
            table_size = format_size(table["data_length"] + table["index_length"])
            for column in table["columns"]:
                rows.append(
                    [
                        schema,
                        module_lookup.get((schema, table["table_name"]), infer_module_name(table["table_name"])),
                        table["table_name"],
                        column["position"],
                        column["name"],
                        column["type"],
                        column["nullable"],
                        column["default"],
                        column["key"] or "",
                        column["extra"] or "",
                        FIELD_MEANINGS.get(column["name"], ""),
                        column["comment"] or "",
                        table["engine"],
                        table["table_rows"],
                        table_size,
                        table["comment"] or "",
                    ]
                )
    return rows


def build_index_rows(
    metadata: dict[str, object],
    module_lookup: dict[tuple[str, str], str],
) -> list[list[object]]:
    tables: dict[str, dict[str, dict[str, object]]] = metadata["tables"]  # type: ignore[assignment]
    rows: list[list[object]] = []
    for schema in metadata["schemas"]:
        for table in sorted(tables[schema].values(), key=lambda item: item["table_name"]):
            for index_name, index_columns in sorted(table["indexes"].items()):
                ordered = sorted(index_columns, key=lambda item: item["seq_in_index"])
                column_desc = []
                for item in ordered:
                    text = item["column_name"]
                    if item["sub_part"]:
                        text += f"({item['sub_part']})"
                    if item["collation"] == "D":
                        text += " DESC"
                    column_desc.append(text)
                rows.append(
                    [
                        schema,
                        module_lookup.get((schema, table["table_name"]), infer_module_name(table["table_name"])),
                        table["table_name"],
                        index_name,
                        ordered[0]["index_type"],
                        "唯一" if ordered[0]["non_unique"] == 0 else "非唯一",
                        len(ordered),
                        ", ".join(column_desc),
                    ]
                )
    return rows


def build_foreign_key_rows(
    metadata: dict[str, object],
    module_lookup: dict[tuple[str, str], str],
) -> list[list[object]]:
    tables: dict[str, dict[str, dict[str, object]]] = metadata["tables"]  # type: ignore[assignment]
    rows: list[list[object]] = []
    for schema in metadata["schemas"]:
        for table in sorted(tables[schema].values(), key=lambda item: item["table_name"]):
            if not table["foreign_keys"]:
                continue
            for item in table["foreign_keys"]:
                rows.append(
                    [
                        schema,
                        module_lookup.get((schema, table["table_name"]), infer_module_name(table["table_name"])),
                        table["table_name"],
                        item["constraint_name"],
                        item["column_name"],
                        item["referenced_table_name"],
                        item["referenced_column_name"],
                    ]
                )
    return rows


def build_module_sheet_rows(module_rows: list[dict[str, object]]) -> list[list[object]]:
    rows: list[list[object]] = []
    for module in module_rows:
        spec = module["spec"]
        rows.append(
            [
                spec["name"],
                spec["schema"],
                len(module["tables"]),
                module["total_rows"],
                format_size(module["total_size"]),
                ", ".join(table["table_name"] for table in module["core_tables"]),
                summarize_table_list([name for name in module["table_names"] if name not in [table["table_name"] for table in module["core_tables"]]], limit=20),
                spec["description"],
                "；".join(spec.get("remarks", [])),
            ]
        )
    return rows


def write_excel_dictionary(metadata: dict[str, object], summaries: dict[str, dict[str, object]]) -> None:
    module_rows = collect_module_rows(metadata)
    module_lookup = build_module_lookup(module_rows)
    wb = Workbook()

    ws_intro = wb.active
    ws_intro.title = "说明"
    intro_rows = [
        ["文档名称", "CRM数据库字段字典"],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["数据源", "数据库.md 中提供的只读 MySQL 连接信息"],
        ["业务库数量", len(metadata["schemas"])],
        ["工作表说明", "库总览 / 表清单 / 字段字典 / 索引字典 / 外键字典 / 模块视图"],
        ["安全说明", "未在工作簿中保存数据库明文密码"],
    ]
    for row in intro_rows:
        ws_intro.append(row)
    apply_header_style(ws_intro, 1, SUBHEADER_FILL, SUBHEADER_FONT)
    style_sheet(ws_intro, freeze_cell="A2", autofilter=False)
    autosize_columns(ws_intro, max_width=80)

    ws_overview = wb.create_sheet("库总览")
    append_sheet_rows(
        ws_overview,
        ["数据库", "表数量", "字段数量", "索引数量", "外键数量", "预计总容量", "无表注释占比"],
        build_overview_rows(metadata, summaries),
    )

    ws_tables = wb.create_sheet("表清单")
    append_sheet_rows(
        ws_tables,
        ["数据库", "业务模块", "表名", "引擎", "字段数", "索引数", "外键数", "预估行数", "数据容量", "索引容量", "总容量", "创建时间", "更新时间", "表注释"],
        build_table_rows(metadata, module_lookup),
    )

    ws_columns = wb.create_sheet("字段字典")
    append_sheet_rows(
        ws_columns,
        ["数据库", "业务模块", "表名", "序号", "字段名", "类型", "可空", "默认值", "键", "额外属性", "字段业务含义", "字段注释", "引擎", "预估行数", "表容量", "表注释"],
        build_column_rows(metadata, module_lookup),
    )

    ws_indexes = wb.create_sheet("索引字典")
    append_sheet_rows(
        ws_indexes,
        ["数据库", "业务模块", "表名", "索引名", "索引类型", "唯一性", "列数量", "索引字段"],
        build_index_rows(metadata, module_lookup),
    )

    ws_foreign_keys = wb.create_sheet("外键字典")
    fk_rows = build_foreign_key_rows(metadata, module_lookup)
    if not fk_rows:
        fk_rows = [["-", "-", "-", "-", "-", "-", "-"]]
    append_sheet_rows(
        ws_foreign_keys,
        ["数据库", "业务模块", "表名", "约束名", "字段", "引用表", "引用字段"],
        fk_rows,
    )

    ws_modules = wb.create_sheet("模块视图")
    append_sheet_rows(
        ws_modules,
        ["模块", "所属库", "表数量", "预估行数", "预估容量", "核心表", "附属/映射表", "模块说明", "备注"],
        build_module_sheet_rows(module_rows),
    )

    wb.save(XLSX_FILE)


def build_overview_markdown(metadata: dict[str, object], summaries: dict[str, dict[str, object]]) -> str:
    schemas: list[str] = metadata["schemas"]  # type: ignore[assignment]
    tables: dict[str, dict[str, dict[str, object]]] = metadata["tables"]  # type: ignore[assignment]
    lines: list[str] = []
    lines.append("# CRM数据库结构分析文档")
    lines.append("")
    lines.append("## 1. 文档说明")
    lines.append("")
    lines.append("- 数据源：`数据库.md` 中提供的只读 MySQL 连接信息。")
    lines.append("- 分析范围：排除系统库后，实际业务库共 2 个。")
    lines.append("- 输出方式：本文档负责总览分析，逐表字段明细见 `docs/db/` 目录。")
    lines.append("- 安全说明：文档中不落地保存数据库明文密码。")
    lines.append("")
    lines.append("## 2. 数据库总览")
    lines.append("")
    lines.append("| 数据库 | 表数量 | 字段数量 | 索引数量 | 外键数量 | 预计总容量 | 无表注释占比 |")
    lines.append("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for schema in schemas:
        summary = summaries[schema]
        no_comment_ratio = 0.0
        if summary["table_count"]:
            no_comment_ratio = summary["no_comment_tables"] / summary["table_count"] * 100
        lines.append(
            f"| `{schema}` | {summary['table_count']} | {summary['column_count']} | {summary['indexes_count']} | "
            f"{summary['foreign_key_count']} | {format_size(summary['total_size'])} | {no_comment_ratio:.1f}% |"
        )
    lines.append("")
    lines.append("## 3. 结构分析结论")
    lines.append("")
    lines.append("### 3.1 `ikcrm_cms_production`")
    lines.append("")
    lines.append("- 该库只有 29 张表，体量很小，更像 CRM 外围的渠道/CMS/代理管理库，而不是主交易库。")
    lines.append("- 典型表包括 `agents`、`agent_users`、`clients`、`risk_packages`、`sms_packages`、`upgrade_notices`。")
    lines.append("- 业务重点偏向代理、客户套餐、通知和试用申请，适合承载官网/渠道侧配置。")
    lines.append("- 表行数整体较小，说明该库以配置类和关系映射类数据为主。")
    lines.append("")
    lines.append("### 3.2 `vcooline_ikcrm_production`")
    lines.append("")
    lines.append("- 该库有 297 张表，是 CRM 主业务库，覆盖线索、客户、联系人、商机、合同、订单、回款、发票、费用、审批、知识库、通讯和生态集成。")
    lines.append("- 存在大量按业务域拆分的资产表、附属表和映射表，如 `customer_assets`、`contract_assets`、`roles_users`、`users_departments`。")
    lines.append("- 集成面较广，包含钉钉、企业微信、King、阿里等生态表，平台属性明显。")
    lines.append("- 大表主要集中在日志、活动、资产和通知类表，运维和容量治理应优先关注这些热点对象。")
    lines.append("")
    lines.append("### 3.3 设计特征")
    lines.append("")
    lines.append("- 绝大多数表没有表注释，字段注释也较少，说明数据库自解释能力弱，后续维护更依赖应用代码和业务口径。")
    lines.append("- 只有极少数外键约束，属于典型的“应用层保证关联关系”的互联网业务库设计。优点是写入灵活，缺点是数据一致性依赖业务代码。")
    lines.append("- 命名总体采用复数英文表名，辅以 `_assets`、`_maps`、`_logs`、`_tracks`、`_reports` 等后缀，具备一定规则性。")
    lines.append("- 多个业务实体采用主表 + 扩展表 + 审批表 + 通知映射表的组合模式，表明系统对审批流、提醒和附加字段扩展较重。")
    lines.append("")
    lines.append("## 4. 模块分布")
    lines.append("")
    for schema in schemas:
        summary = summaries[schema]
        lines.append(f"### 4.{schemas.index(schema) + 1} `{schema}`")
        lines.append("")
        lines.append("| 模块 | 表数量 |")
        lines.append("| --- | ---: |")
        for module, count in summary["modules"].most_common():
            lines.append(f"| {module} | {count} |")
        lines.append("")
    lines.append("## 5. 大表与容量关注点")
    lines.append("")
    for schema in schemas:
        summary = summaries[schema]
        lines.append(f"### 5.{schemas.index(schema) + 1} `{schema}`")
        lines.append("")
        lines.append("| 表名 | 预估行数 | 预估容量 |")
        lines.append("| --- | ---: | ---: |")
        for item in summary["largest_tables"]:
            lines.append(f"| `{item['table_name']}` | {item['rows']} | {format_size(item['size_bytes'])} |")
        lines.append("")
    lines.append("## 6. 风险与建议")
    lines.append("")
    lines.append("1. 元数据治理偏弱。建议补齐核心表和关键字段注释，至少覆盖客户、商机、合同、订单、回款、权限相关表。")
    lines.append("2. 主业务库存在较多超大日志/资产表。建议结合分区、归档和冷热分层策略优化 `operation_logs`、`token_logs`、`sales_activities` 等对象。")
    lines.append("3. 外键约束很少，建议在应用侧补充关联完整性校验，并为关键映射表建立巡检 SQL。")
    lines.append("4. 命名虽有规律，但仍存在并行命名，如 `callagents` 与 `call_agents`、`callcenters` 与 `call_centers`，建议评估历史兼容表并整理标准。")
    lines.append("5. 由于 `ikcrm_cms_production` 和 `vcooline_ikcrm_production` 分别承担外围渠道与核心 CRM 职责，建议后续文档和开发也按此边界维护。")
    lines.append("")
    lines.append("## 7. 明细文档")
    lines.append("")
    for schema in schemas:
        lines.append(f"- `{schema}`：`docs/db/{schema}.md`")
    lines.append("")
    lines.append("## 8. 重点表抽样")
    lines.append("")
    key_tables = [
        ("ikcrm_cms_production", "agent_users"),
        ("vcooline_ikcrm_production", "customers"),
        ("vcooline_ikcrm_production", "opportunities"),
        ("vcooline_ikcrm_production", "contracts"),
        ("vcooline_ikcrm_production", "users"),
    ]
    for schema, table_name in key_tables:
        table = tables.get(schema, {}).get(table_name)
        if not table:
            continue
        lines.append(f"### {schema}.{table_name}")
        lines.append("")
        lines.append(
            f"- 引擎：`{table['engine']}`；字段数：{len(table['columns'])}；索引数：{len(table['indexes'])}；"
            f" 预估行数：{table['table_rows']}；预估容量：{format_size(table['data_length'] + table['index_length'])}。"
        )
        sample_columns = ", ".join(column["name"] for column in table["columns"][:8])
        lines.append(f"- 前 8 个字段：`{sample_columns}`。")
        if table["foreign_keys"]:
            fk_desc = ", ".join(
                f"`{item['column_name']}` -> `{item['referenced_table_name']}.{item['referenced_column_name']}`"
                for item in table["foreign_keys"]
            )
            lines.append(f"- 外键：{fk_desc}")
        else:
            lines.append("- 外键：未在 `information_schema` 中发现显式外键约束。")
        lines.append("")
    return "\n".join(lines) + "\n"


def render_table_detail(table: dict[str, object]) -> str:
    lines: list[str] = []
    table_size = table["data_length"] + table["index_length"]
    lines.append(f"## {table['table_name']}")
    lines.append("")
    lines.append("| 属性 | 值 |")
    lines.append("| --- | --- |")
    lines.append(f"| 存储引擎 | `{table['engine']}` |")
    lines.append(f"| 预估行数 | {table['table_rows']} |")
    lines.append(f"| 数据容量 | {format_size(table['data_length'])} |")
    lines.append(f"| 索引容量 | {format_size(table['index_length'])} |")
    lines.append(f"| 总容量 | {format_size(table_size)} |")
    lines.append(f"| 创建时间 | {table['create_time'] or '-'} |")
    lines.append(f"| 更新时间 | {table['update_time'] or '-'} |")
    lines.append(f"| 表注释 | {table['comment'] or '-'} |")
    lines.append("")
    lines.append("### 字段")
    lines.append("")
    lines.append("| 序号 | 字段名 | 类型 | 可空 | 默认值 | 键 | 额外属性 | 注释 |")
    lines.append("| ---: | --- | --- | --- | --- | --- | --- | --- |")
    for column in table["columns"]:
        default_value = column["default"]
        if default_value == "NULL":
            default_value = "NULL"
        default_value = str(default_value).replace("\n", " ").replace("|", "\\|")
        comment = str(column["comment"]).replace("\n", " ").replace("|", "\\|") or "-"
        extra = str(column["extra"]).replace("\n", " ").replace("|", "\\|") or "-"
        key = column["key"] or "-"
        lines.append(
            f"| {column['position']} | `{column['name']}` | `{column['type']}` | {column['nullable']} | "
            f"{default_value} | {key} | {extra} | {comment} |"
        )
    lines.append("")
    lines.append("### 索引")
    lines.append("")
    lines.append("| 索引名 | 类型 | 唯一性 | 列 |")
    lines.append("| --- | --- | --- | --- |")
    if table["indexes"]:
        for index_name, columns in sorted(table["indexes"].items()):
            ordered_columns = sorted(columns, key=lambda item: item["seq_in_index"])
            column_desc = []
            for item in ordered_columns:
                col = f"`{item['column_name']}`"
                if item["sub_part"]:
                    col += f"({item['sub_part']})"
                if item["collation"] == "D":
                    col += " DESC"
                column_desc.append(col)
            unique_text = "唯一" if ordered_columns[0]["non_unique"] == 0 else "非唯一"
            lines.append(
                f"| `{index_name}` | {ordered_columns[0]['index_type']} | {unique_text} | {', '.join(column_desc)} |"
            )
    else:
        lines.append("| - | - | - | - |")
    lines.append("")
    lines.append("### 外键")
    lines.append("")
    lines.append("| 约束名 | 字段 | 引用表 | 引用字段 |")
    lines.append("| --- | --- | --- | --- |")
    if table["foreign_keys"]:
        for item in table["foreign_keys"]:
            lines.append(
                f"| `{item['constraint_name']}` | `{item['column_name']}` | "
                f"`{item['referenced_table_name']}` | `{item['referenced_column_name']}` |"
            )
    else:
        lines.append("| - | - | - | - |")
    lines.append("")
    return "\n".join(lines)


def build_schema_markdown(schema: str, schema_tables: dict[str, dict[str, object]], summary: dict[str, object]) -> str:
    lines: list[str] = []
    lines.append(f"# {schema} 表结构明细")
    lines.append("")
    lines.append("## 1. 数据库概况")
    lines.append("")
    lines.append("| 指标 | 值 |")
    lines.append("| --- | ---: |")
    lines.append(f"| 表数量 | {summary['table_count']} |")
    lines.append(f"| 字段数量 | {summary['column_count']} |")
    lines.append(f"| 索引数量 | {summary['indexes_count']} |")
    lines.append(f"| 外键数量 | {summary['foreign_key_count']} |")
    lines.append(f"| 预计总容量 | {format_size(summary['total_size'])} |")
    lines.append("")
    lines.append("## 2. 表清单")
    lines.append("")
    lines.append("| 表名 | 字段数 | 索引数 | 外键数 | 预估行数 | 预估容量 | 表注释 |")
    lines.append("| --- | ---: | ---: | ---: | ---: | ---: | --- |")
    sorted_tables = sorted(schema_tables.values(), key=lambda item: item["table_name"])
    for table in sorted_tables:
        comment = str(table["comment"]).replace("|", "\\|") or "-"
        lines.append(
            f"| `{table['table_name']}` | {len(table['columns'])} | {len(table['indexes'])} | "
            f"{len(table['foreign_keys'])} | {table['table_rows']} | "
            f"{format_size(table['data_length'] + table['index_length'])} | {comment or '-'} |"
        )
    lines.append("")
    lines.append("## 3. 逐表详情")
    lines.append("")
    for table in sorted_tables:
        lines.append(render_table_detail(table))
    return "\n".join(lines) + "\n"


def ensure_output_dirs() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def main() -> None:
    ensure_output_dirs()
    config = parse_db_config(SOURCE_FILE)
    metadata = load_metadata(config)
    summaries = collect_summary(metadata)
    OVERVIEW_FILE.write_text(build_overview_markdown(metadata, summaries), encoding="utf-8")
    MODULE_FILE.write_text(build_module_markdown(metadata, summaries), encoding="utf-8")
    RELATION_FILE.write_text(build_relation_markdown(metadata), encoding="utf-8")
    write_excel_dictionary(metadata, summaries)
    for schema in metadata["schemas"]:
        schema_file = OUTPUT_DIR / f"{schema}.md"
        schema_file.write_text(
            build_schema_markdown(schema, metadata["tables"][schema], summaries[schema]),
            encoding="utf-8",
        )
    print(f"Generated overview: {OVERVIEW_FILE}")
    print(f"Generated module view: {MODULE_FILE}")
    print(f"Generated relation view: {RELATION_FILE}")
    print(f"Generated excel dictionary: {XLSX_FILE}")
    for schema in metadata["schemas"]:
        print(f"Generated detail: {OUTPUT_DIR / f'{schema}.md'}")


if __name__ == "__main__":
    main()
